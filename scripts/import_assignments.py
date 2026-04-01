#!/usr/bin/env python3
"""
Import evaluation assignments from Excel file.
Format: Sheet2 has evaluatee per row with columns:
  A: รหัสพนักงาน (evaluatee emid)
  B-F: ข้อมูลพนักงาน
  J: ประเมินตนเอง (/ = yes)
  K: องศาบน (evaluator names, comma/newline separated)
  L: องศาล่าง
  M: องศาซ้าย

Usage:
  python scripts/import_assignments.py <excel_file> <fiscal_year> [--dry-run]

Example:
  python scripts/import_assignments.py "docs/360 ปี 2568 ระดับ 4 (11 คน).xlsx" 2025 --dry-run
  python scripts/import_assignments.py "docs/360 ปี 2568 ระดับ 4 (11 คน).xlsx" 2025
"""

import sys
import os
import re
import json
import openpyxl

def parse_names(cell_value):
    """Parse evaluator names from a cell (comma or newline separated)."""
    if not cell_value or cell_value.strip() in ('ไม่มี', '-', ''):
        return []

    text = str(cell_value).strip()
    # Split by newline first, then by comma
    parts = []
    for line in text.split('\n'):
        for part in line.split(','):
            name = part.strip()
            if name and name != 'ไม่มี':
                parts.append(name)
    return parts

def normalize_name(name):
    """Normalize Thai name: remove prefix, extra spaces."""
    name = re.sub(r'\s+', ' ', name).strip()
    # Remove common prefixes for matching
    prefixes = ['นาย', 'นาง', 'นางสาว', 'ว่าที่ร้อยตรี', 'ดร.']
    clean = name
    for p in prefixes:
        if clean.startswith(p + ' '):
            clean = clean[len(p):].strip()
        elif clean.startswith(p):
            clean = clean[len(p):].strip()
    return clean

def generate_sql(excel_path, fiscal_year, dry_run=False):
    """Generate SQL INSERT statements from Excel."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    if 'Sheet2' not in wb.sheetnames:
        print("ERROR: Sheet2 not found in Excel file")
        sys.exit(1)

    ws = wb['Sheet2']

    assignments = []
    unmatched_names = set()

    for row in ws.iter_rows(min_row=2, values_only=False):
        emid = row[0].value  # A: รหัสพนักงาน
        if not emid:
            continue

        evaluatee_emid = str(emid).strip()
        grade = row[5].value  # F: ระดับ
        self_eval = str(row[9].value or '').strip()  # J: ประเมินตนเอง
        top_names = parse_names(row[10].value)  # K: องศาบน
        bottom_names = parse_names(row[11].value)  # L: องศาล่าง
        left_names = parse_names(row[12].value)  # M: องศาซ้าย

        # Self evaluation
        if self_eval == '/':
            assignments.append({
                'evaluatee_emid': evaluatee_emid,
                'evaluator_emid': evaluatee_emid,  # self
                'angle': 'self',
            })

        # Top (องศาบน)
        for name in top_names:
            assignments.append({
                'evaluatee_emid': evaluatee_emid,
                'evaluator_name': name,
                'angle': 'top',
            })

        # Bottom (องศาล่าง)
        for name in bottom_names:
            assignments.append({
                'evaluatee_emid': evaluatee_emid,
                'evaluator_name': name,
                'angle': 'bottom',
            })

        # Left (องศาซ้าย)
        for name in left_names:
            assignments.append({
                'evaluatee_emid': evaluatee_emid,
                'evaluator_name': name,
                'angle': 'left',
            })

    print(f"Parsed {len(assignments)} assignments from Excel")
    print(f"Fiscal year: {fiscal_year}")
    print()

    # Generate PHP artisan tinker script (more reliable than raw SQL for name matching)
    php_lines = []
    php_lines.append(f'$fiscalYear = {fiscal_year};')
    php_lines.append('$created = 0; $skipped = 0; $notFound = 0; $errors = [];')
    php_lines.append('')
    php_lines.append('// Build user lookup by emid and by name')
    php_lines.append('$allUsers = \\App\\Models\\User::all();')
    php_lines.append('$byEmid = $allUsers->keyBy("emid");')
    php_lines.append('$byName = [];')
    php_lines.append('foreach ($allUsers as $u) {')
    php_lines.append('    $key = trim($u->prename . $u->fname . " " . $u->lname);')
    php_lines.append('    $byName[$key] = $u;')
    php_lines.append('    // Also without space after prename')
    php_lines.append('    $key2 = trim($u->fname . " " . $u->lname);')
    php_lines.append('    if (!isset($byName[$key2])) $byName[$key2] = $u;')
    php_lines.append('}')
    php_lines.append('')
    php_lines.append('function findUser($nameOrEmid, $byEmid, $byName) {')
    php_lines.append('    $nameOrEmid = trim($nameOrEmid);')
    php_lines.append('    if (isset($byEmid[$nameOrEmid])) return $byEmid[$nameOrEmid];')
    php_lines.append('    if (isset($byName[$nameOrEmid])) return $byName[$nameOrEmid];')
    php_lines.append('    // Try fuzzy: remove prefix')
    php_lines.append('    $clean = preg_replace("/^(นาย|นาง|นางสาว|ว่าที่ร้อยตรี|ดร\\.)\\s*/u", "", $nameOrEmid);')
    php_lines.append('    foreach ($byName as $k => $v) {')
    php_lines.append('        if (str_contains($k, $clean)) return $v;')
    php_lines.append('    }')
    php_lines.append('    return null;')
    php_lines.append('}')
    php_lines.append('')

    # Use EvaluationLookupService for correct eval
    php_lines.append('$assignments = [')
    for a in assignments:
        if 'evaluator_emid' in a:
            php_lines.append(f'    ["evaluatee" => "{a["evaluatee_emid"]}", "evaluator" => "{a["evaluator_emid"]}", "angle" => "{a["angle"]}"],')
        else:
            name_escaped = a['evaluator_name'].replace('"', '\\"')
            php_lines.append(f'    ["evaluatee" => "{a["evaluatee_emid"]}", "evaluator_name" => "{name_escaped}", "angle" => "{a["angle"]}"],')
    php_lines.append('];')
    php_lines.append('')

    if dry_run:
        php_lines.append('$dryRun = true;')
    else:
        php_lines.append('$dryRun = false;')

    php_lines.append('')
    php_lines.append('foreach ($assignments as $a) {')
    php_lines.append('    $evaluatee = $byEmid[$a["evaluatee"]] ?? null;')
    php_lines.append('    if (!$evaluatee) { $errors[] = "Evaluatee not found: " . $a["evaluatee"]; $notFound++; continue; }')
    php_lines.append('')
    php_lines.append('    if (isset($a["evaluator"])) {')
    php_lines.append('        $evaluator = $byEmid[$a["evaluator"]] ?? null;')
    php_lines.append('    } else {')
    php_lines.append('        $evaluator = findUser($a["evaluator_name"], $byEmid, $byName);')
    php_lines.append('    }')
    php_lines.append('    if (!$evaluator) { $errors[] = "Evaluator not found: " . ($a["evaluator_name"] ?? $a["evaluator"]); $notFound++; continue; }')
    php_lines.append('')
    php_lines.append('    // Find correct evaluation form')
    php_lines.append('    $grade = (int) $evaluatee->grade;')
    php_lines.append('    $userType = $evaluatee->user_type instanceof \\BackedEnum ? $evaluatee->user_type->value : ($evaluatee->user_type ?? "internal");')
    php_lines.append('    $eval = \\App\\Services\\EvaluationLookupService::findByGrade($grade, $userType, $fiscalYear);')
    php_lines.append('    if (!$eval) { $errors[] = "No eval for grade=$grade fy=$fiscalYear"; $notFound++; continue; }')
    php_lines.append('')
    php_lines.append('    // Check duplicate')
    php_lines.append('    $exists = \\App\\Models\\EvaluationAssignment::where("evaluator_id", $evaluator->id)')
    php_lines.append('        ->where("evaluatee_id", $evaluatee->id)')
    php_lines.append('        ->where("fiscal_year", $fiscalYear)')
    php_lines.append('        ->where("angle", $a["angle"])')
    php_lines.append('        ->exists();')
    php_lines.append('    if ($exists) { $skipped++; continue; }')
    php_lines.append('')
    php_lines.append('    if (!$dryRun) {')
    php_lines.append('        \\App\\Models\\EvaluationAssignment::create([')
    php_lines.append('            "evaluation_id" => $eval->id,')
    php_lines.append('            "evaluator_id" => $evaluator->id,')
    php_lines.append('            "evaluatee_id" => $evaluatee->id,')
    php_lines.append('            "fiscal_year" => $fiscalYear,')
    php_lines.append('            "angle" => $a["angle"],')
    php_lines.append('        ]);')
    php_lines.append('    }')
    php_lines.append('    $created++;')
    php_lines.append('}')
    php_lines.append('')
    php_lines.append('echo "\\n=== Result ===\\n";')
    php_lines.append('echo ($dryRun ? "[DRY RUN] " : "") . "Created: $created | Skipped (dup): $skipped | Not found: $notFound\\n";')
    php_lines.append('if ($errors) { echo "\\nErrors:\\n"; foreach(array_unique($errors) as $e) echo "  - $e\\n"; }')

    return '\n'.join(php_lines)


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Usage: python scripts/import_assignments.py <excel_file> <fiscal_year> [--dry-run]")
        print('Example: python scripts/import_assignments.py "docs/360 ปี 2568 ระดับ 4 (11 คน).xlsx" 2025 --dry-run')
        sys.exit(1)

    excel_path = sys.argv[1]
    fiscal_year = int(sys.argv[2])
    dry_run = '--dry-run' in sys.argv

    if not os.path.exists(excel_path):
        print(f"ERROR: File not found: {excel_path}")
        sys.exit(1)

    php_script = generate_sql(excel_path, fiscal_year, dry_run)

    output_path = 'scripts/import_assignments_generated.php'
    os.makedirs('scripts', exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(php_script)

    print(f"Generated PHP script: {output_path}")
    print(f"{'[DRY RUN]' if dry_run else '[LIVE]'}")
    print()
    print("To run on server:")
    print(f'  php artisan tinker --execute="$(cat {output_path})"')
    print()
    print("Or copy to server and run:")
    print(f'  scp {output_path} server:/tmp/')
    print(f'  ssh server "cd laravel_project && php artisan tinker < /tmp/import_assignments_generated.php"')
